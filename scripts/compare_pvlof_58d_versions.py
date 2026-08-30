"""Build a 58-day multi-version PVLOF comparison package.

The input root must contain plant_id=<id> directories created by
run_pvlof_58d_versions.py. Outputs are compact CSV/JSON review tables; the
multi-million-row point Parquets remain under artifacts/models.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = PROJECT_ROOT / "src"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))


import pandas as pd


DEVICE_TIME_KEYS = ["plant_id", "device_no", "event_time"]
VERSION_SPECS = [
    {
        "key": "v1_5",
        "label": "PVLOF_V1_5_MEMORY_5PCT",
        "points": "pvlof_v15_full_points.parquet",
        "evidence": "pvlof_v15_evidence_points.parquet",
        "raw": "pvlof_v15_raw_anomaly",
        "final": "pvlof_v15_alert",
    },
    {
        "key": "v1_6",
        "label": "PVLOF_V1_6_HYBRID_GATE",
        "points": "pvlof_v16_full_points.parquet",
        "evidence": "pvlof_v16_evidence_points.parquet",
        "raw": "pvlof_v16_raw_anomaly",
        "final": "pvlof_v16_alert",
    },
    {
        "key": "v1_7",
        "label": "PVLOF_V1_7_PENALIZED_SEGMENTATION",
        "points": "pvlof_v17_full_points.parquet",
        "evidence": "pvlof_v17_evidence_points.parquet",
        "raw": "pvlof_v17_raw_anomaly",
        "final": "pvlof_v17_alert",
    },
    {
        "key": "v1_7_improved",
        "label": "PVLOF_V1_7_IMPROVED",
        "points": "pvlof_v17_improved_full_points.parquet",
        "evidence": "pvlof_v17_improved_evidence_points.parquet",
        "raw": "pvlof_v17_improved_raw_anomaly",
        "final": "pvlof_v17_improved_alert",
    },
    {
        "key": "v1_7_improved_v2",
        "label": "PVLOF_V1_7_IMPROVED_V2",
        "points": "pvlof_v17_improved_v2_full_points.parquet",
        "evidence": "pvlof_v17_improved_v2_evidence_points.parquet",
        "raw": "pvlof_v17_improved_raw_anomaly",
        "final": "pvlof_v17_improved_alert",
    },
]
VERSION_LABELS = [spec["label"] for spec in VERSION_SPECS]
IMPROVED_LABEL = "PVLOF_V1_7_IMPROVED"
IMPROVED_V2_LABEL = "PVLOF_V1_7_IMPROVED_V2"


def _numbers(values: pd.Series) -> str:
    numbers: set[int] = set()
    for value in values.dropna():
        for item in str(value).split(","):
            if item.strip():
                numbers.add(int(float(item)))
    return ",".join(f"{number:02d}" for number in sorted(numbers))


def _number_set(value: Any) -> set[int]:
    if pd.isna(value) or not str(value).strip():
        return set()
    return {int(float(item)) for item in str(value).split(",") if item.strip()}


def _format_numbers(values: set[int]) -> str:
    return ",".join(f"{number:02d}" for number in sorted(values))


def _read_signal(path: Path, column: str, label: str) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    columns = [*DEVICE_TIME_KEYS, "string_no", column]
    frame = pd.read_parquet(path, columns=columns)
    frame["plant_id"] = frame["plant_id"].astype(str)
    frame["device_no"] = frame["device_no"].astype(str)
    frame["event_time"] = pd.to_datetime(frame["event_time"], errors="raise", utc=True)
    selected = frame[frame[column].fillna(False).astype(bool)]
    if selected.empty:
        return pd.DataFrame(columns=[*DEVICE_TIME_KEYS, label])
    return (
        selected.groupby(DEVICE_TIME_KEYS, observed=True)["string_no"]
        .agg(_numbers)
        .rename(label)
        .reset_index()
    )


def _read_confirmed_evidence(path: Path, label: str) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    frame = pd.read_parquet(path)
    required = {*DEVICE_TIME_KEYS, "string_no", "raw_evidence", "final_alert"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"Evidence file {path} is missing columns: {missing}")
    frame["plant_id"] = frame["plant_id"].astype(str)
    frame["device_no"] = frame["device_no"].astype(str)
    frame["event_time"] = pd.to_datetime(frame["event_time"], errors="raise", utc=True)
    selected = frame[
        frame["raw_evidence"].fillna(False).astype(bool)
        | frame["final_alert"].fillna(False).astype(bool)
    ]
    if selected.empty:
        return pd.DataFrame(columns=[*DEVICE_TIME_KEYS, label])
    return (
        selected.groupby(DEVICE_TIME_KEYS, observed=True)["string_no"]
        .agg(_numbers)
        .rename(label)
        .reset_index()
    )


def _merge_version_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    result: pd.DataFrame | None = None
    for frame in frames:
        result = (
            frame.copy()
            if result is None
            else result.merge(frame, on=DEVICE_TIME_KEYS, how="outer", validate="one_to_one")
        )
    if result is None:
        return pd.DataFrame(columns=[*DEVICE_TIME_KEYS, *VERSION_LABELS])
    for label in VERSION_LABELS:
        if label not in result:
            result[label] = ""
    result[VERSION_LABELS] = result[VERSION_LABELS].astype("string").fillna("")
    return result.sort_values(DEVICE_TIME_KEYS).reset_index(drop=True)


def _comparison_case(row: pd.Series) -> str:
    present = [spec for spec in VERSION_SPECS if str(row[spec["label"]])]
    if len(present) == 1:
        return f"only_{present[0]['key']}"
    values = {str(row[spec["label"]]) for spec in present}
    if len(present) == len(VERSION_SPECS) and len(values) == 1:
        return "all_same"
    if len(values) == 1:
        return "same_members_partial_versions"
    return "different_members"


def _v2_case(row: pd.Series) -> str:
    old = str(row[IMPROVED_LABEL])
    new = str(row[IMPROVED_V2_LABEL])
    if old and new:
        return "both_same" if old == new else "both_different"
    if old:
        return "improved_only"
    if new:
        return "v2_only"
    return "neither"


def add_comparison_columns(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result["present_versions"] = result.apply(
        lambda row: ",".join(
            spec["key"] for spec in VERSION_SPECS if str(row[spec["label"]])
        ),
        axis=1,
    )
    result["version_count"] = result["present_versions"].map(
        lambda value: 0 if not value else len(str(value).split(","))
    )
    result["comparison_case"] = result.apply(_comparison_case, axis=1)
    result["v2_vs_improved_case"] = result.apply(_v2_case, axis=1)
    result["v2_added_strings"] = result.apply(
        lambda row: _format_numbers(
            _number_set(row[IMPROVED_V2_LABEL]) - _number_set(row[IMPROVED_LABEL])
        ),
        axis=1,
    )
    result["v2_removed_strings"] = result.apply(
        lambda row: _format_numbers(
            _number_set(row[IMPROVED_LABEL]) - _number_set(row[IMPROVED_V2_LABEL])
        ),
        axis=1,
    )
    return result


def _read_layer(
    plant_directories: list[Path],
    *,
    signal: str,
) -> pd.DataFrame:
    per_plant: list[pd.DataFrame] = []
    for plant_directory in plant_directories:
        frames = [
            _read_signal(
                plant_directory / spec["points"],
                str(spec[signal]),
                str(spec["label"]),
            )
            for spec in VERSION_SPECS
        ]
        per_plant.append(_merge_version_frames(frames))
    if not per_plant:
        return pd.DataFrame(columns=[*DEVICE_TIME_KEYS, *VERSION_LABELS])
    return add_comparison_columns(
        pd.concat(per_plant, ignore_index=True).sort_values(DEVICE_TIME_KEYS)
    ).reset_index(drop=True)


def _read_confirmed_candidate_points(plant_directories: list[Path]) -> pd.DataFrame:
    per_plant: list[pd.DataFrame] = []
    for plant_directory in plant_directories:
        frames = [
            _read_confirmed_evidence(
                plant_directory / spec["evidence"], str(spec["label"])
            )
            for spec in VERSION_SPECS
        ]
        per_plant.append(_merge_version_frames(frames))
    if not per_plant:
        return pd.DataFrame(columns=[*DEVICE_TIME_KEYS, *VERSION_LABELS])
    return add_comparison_columns(
        pd.concat(per_plant, ignore_index=True).sort_values(DEVICE_TIME_KEYS)
    ).reset_index(drop=True)


def build_confirmed_events(
    points: pd.DataFrame,
    *,
    timezone: str,
    interval_minutes: int,
) -> pd.DataFrame:
    if points.empty:
        return pd.DataFrame(
            columns=[
                "row",
                "event_id",
                "plant_id",
                "device_no",
                "raise_time_local",
                "end_time_local",
                *VERSION_LABELS,
                "comparison_case",
                "v2_vs_improved_case",
                "v2_added_strings",
                "v2_removed_strings",
                "alert_time_points",
                "duration_minutes",
                "manual_label",
                "manual_strings",
                "review_note",
            ]
        )
    source = points.sort_values(DEVICE_TIME_KEYS).reset_index(drop=True).copy()
    expected = pd.Timedelta(minutes=interval_minutes)
    source["_event"] = (
        source["plant_id"].ne(source["plant_id"].shift())
        | source["device_no"].ne(source["device_no"].shift())
        | source["event_time"].sub(source["event_time"].shift()).ne(expected)
    ).cumsum()
    source["event_id"] = source["_event"].map(
        {
            value: f"pvlof-58d-{number:06d}"
            for number, value in enumerate(source["_event"].unique(), 1)
        }
    )
    aggregations: dict[str, tuple[str, Any]] = {
        "raise_time": ("event_time", "min"),
        "end_time": ("event_time", "max"),
        "alert_time_points": ("event_time", "size"),
    }
    for label in VERSION_LABELS:
        aggregations[label] = (label, _numbers)
    events = (
        source.groupby(
            ["event_id", "plant_id", "device_no", "_event"], observed=True
        )
        .agg(**aggregations)
        .reset_index()
        .drop(columns="_event")
    )
    events = add_comparison_columns(events)
    events["raise_time_local"] = (
        events["raise_time"].dt.tz_convert(timezone).dt.strftime("%Y-%m-%d %H:%M:%S")
    )
    events["end_time_local"] = (
        events["end_time"].dt.tz_convert(timezone).dt.strftime("%Y-%m-%d %H:%M:%S")
    )
    events["duration_minutes"] = (
        (events["end_time"] - events["raise_time"]).dt.total_seconds() / 60
        + interval_minutes
    ).astype(int)
    events["manual_label"] = ""
    events["manual_strings"] = ""
    events["review_note"] = ""
    events = events.sort_values(["plant_id", "device_no", "raise_time"]).reset_index(
        drop=True
    )
    events.insert(0, "row", range(1, len(events) + 1))
    columns = [
        "row",
        "event_id",
        "plant_id",
        "device_no",
        "raise_time_local",
        "end_time_local",
        *VERSION_LABELS,
        "comparison_case",
        "v2_vs_improved_case",
        "v2_added_strings",
        "v2_removed_strings",
        "alert_time_points",
        "duration_minutes",
        "manual_label",
        "manual_strings",
        "review_note",
    ]
    return events[columns]


def _string_count(value: Any) -> int:
    return len(_number_set(value))


def _scope_summary(
    raw_points: pd.DataFrame,
    final_points: pd.DataFrame,
    confirmed_events: pd.DataFrame,
    *,
    keys: list[str],
    universe: pd.DataFrame | None = None,
) -> pd.DataFrame:
    sources = []
    for frame in (raw_points, final_points, confirmed_events):
        if not frame.empty:
            sources.append(frame[keys])
    if universe is not None:
        result = universe.drop_duplicates(keys).reset_index(drop=True)
    elif not sources:
        return pd.DataFrame(columns=keys)
    else:
        result = pd.concat(sources, ignore_index=True).drop_duplicates().reset_index(drop=True)

    def attach_counts(frame: pd.DataFrame, label: str, prefix: str) -> None:
        nonlocal result
        selected = frame[frame[label].astype("string").fillna("").ne("")].copy()
        selected["_string_count"] = selected[label].map(_string_count)
        grouped = selected.groupby(keys, observed=True).agg(
            **{
                f"{prefix}_rows": (label, "size"),
                f"{prefix}_string_points": ("_string_count", "sum"),
            }
        ).reset_index()
        result = result.merge(grouped, on=keys, how="left", validate="one_to_one")

    for spec in VERSION_SPECS:
        key = str(spec["key"])
        label = str(spec["label"])
        attach_counts(raw_points, label, f"raw_{key}")
        attach_counts(final_points, label, f"final_{key}")
        attach_counts(confirmed_events, label, f"events_{key}")

    changed = confirmed_events[
        confirmed_events["v2_vs_improved_case"].isin(
            ["improved_only", "v2_only", "both_different"]
        )
    ]
    changed_counts = (
        changed.groupby(keys, observed=True)
        .size()
        .rename("v2_changed_events")
        .reset_index()
    )
    result = result.merge(changed_counts, on=keys, how="left", validate="one_to_one")
    numeric = [column for column in result.columns if column not in keys]
    count_columns = [column for column in numeric if not column.endswith("_local")]
    result[count_columns] = result[count_columns].fillna(0).astype(int)
    if "device_time_rows" in result:
        denominator = result["device_time_rows"].where(result["device_time_rows"].gt(0))
        for spec in VERSION_SPECS:
            key = str(spec["key"])
            result[f"raw_{key}_rate"] = result[f"raw_{key}_rows"] / denominator
            result[f"final_{key}_rate"] = result[f"final_{key}_rows"] / denominator
    return result.sort_values(keys).reset_index(drop=True)


def _global_version_summary(
    raw_points: pd.DataFrame,
    final_points: pd.DataFrame,
    confirmed_points: pd.DataFrame,
    confirmed_events: pd.DataFrame,
    *,
    evaluated_device_times: int,
) -> dict[str, Any]:
    report: dict[str, Any] = {}
    for spec in VERSION_SPECS:
        key = str(spec["key"])
        label = str(spec["label"])

        def counts(frame: pd.DataFrame) -> tuple[int, int]:
            selected = frame[frame[label].astype("string").fillna("").ne("")]
            return int(len(selected)), int(selected[label].map(_string_count).sum())

        raw_rows, raw_strings = counts(raw_points)
        final_rows, final_strings = counts(final_points)
        confirmed_rows, confirmed_strings = counts(confirmed_points)
        event_rows, event_strings = counts(confirmed_events)
        report[key] = {
            "label": label,
            "raw_candidate_device_times": raw_rows,
            "raw_candidate_string_points": raw_strings,
            "raw_candidate_device_time_rate": (
                raw_rows / evaluated_device_times if evaluated_device_times else 0.0
            ),
            "final_alert_device_times": final_rows,
            "final_alert_string_points": final_strings,
            "final_alert_device_time_rate": (
                final_rows / evaluated_device_times if evaluated_device_times else 0.0
            ),
            "confirmed_candidate_device_times": confirmed_rows,
            "confirmed_candidate_string_points": confirmed_strings,
            "confirmed_events": event_rows,
            "event_member_strings": event_strings,
        }
    return report


def _load_contracts(plant_directories: list[Path]) -> list[dict[str, Any]]:
    contracts = []
    expected: tuple[Any, ...] | None = None
    for directory in plant_directories:
        path = directory / "summary.json"
        if not path.exists():
            raise FileNotFoundError(path)
        payload = json.loads(path.read_text(encoding="utf-8"))
        contract = payload.get("contract", {})
        signature = (
            contract.get("start_local_inclusive"),
            contract.get("end_local_exclusive"),
            contract.get("timezone"),
            contract.get("interval_minutes"),
            contract.get("alarm_device_filter"),
        )
        if expected is None:
            expected = signature
        elif signature != expected:
            raise ValueError(f"Replay contracts differ across plants: {directory}")
        if contract.get("alarm_device_filter") is not False:
            raise ValueError(f"Replay is not full-context: {directory}")
        contracts.append(contract)
    return contracts


def _load_device_coverage(plant_directories: list[Path]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for directory in plant_directories:
        path = directory / "device_coverage.csv"
        if not path.exists():
            raise FileNotFoundError(path)
        frame = pd.read_csv(path, dtype={"plant_id": "string", "device_no": "string"})
        required = {
            "plant_id",
            "device_no",
            "device_time_rows",
            "first_time_local",
            "last_time_local",
        }
        missing = sorted(required - set(frame.columns))
        if missing:
            raise ValueError(f"Coverage file {path} is missing columns: {missing}")
        frames.append(frame)
    return pd.concat(frames, ignore_index=True).drop_duplicates(
        ["plant_id", "device_no"]
    ).sort_values(
        ["plant_id", "device_no"]
    ).reset_index(drop=True)


def _prepare_output(output: Path, *, overwrite: bool) -> None:
    if output.exists() and any(output.iterdir()) and not overwrite:
        raise FileExistsError(
            f"Output directory is not empty: {output}. Use --overwrite to replace "
            "the known comparison files without deleting unrelated files."
        )
    output.mkdir(parents=True, exist_ok=True)


def _summary_rows(summary: dict[str, Any]) -> pd.DataFrame:
    rows = [
        {"section": "scope", "metric": key, "value": value}
        for key, value in summary["scope"].items()
    ]
    rows.append({
        "section": "note",
        "metric": "interpretation",
        "value": "58天无完整人工真值；本报告比较版本行为，不代表准确率。",
    })
    for key, metrics in summary["versions"].items():
        label = metrics["label"]
        for metric, value in metrics.items():
            if metric != "label":
                rows.append({
                    "section": f"version:{key}",
                    "metric": metric,
                    "value": value,
                    "version_label": label,
                })
    for metric, value in summary["comparison"].items():
        if isinstance(value, dict):
            for case, count in value.items():
                rows.append({
                    "section": "comparison",
                    "metric": f"{metric}:{case}",
                    "value": count,
                })
        else:
            rows.append({
                "section": "comparison",
                "metric": metric,
                "value": value,
            })
    return pd.DataFrame(rows, columns=["section", "version_label", "metric", "value"])


def _write_workbook(
    path: Path,
    *,
    summary: dict[str, Any],
    by_plant: pd.DataFrame,
    by_device: pd.DataFrame,
    confirmed_events: pd.DataFrame,
    v2_raw_changes: pd.DataFrame,
    v2_changes: pd.DataFrame,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "Excel output requires openpyxl. The CSV/JSON comparison logic itself "
            "does not require it."
        ) from error

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2F3")
    change_fills = {
        "v2_only": PatternFill("solid", fgColor="E2F0D9"),
        "improved_only": PatternFill("solid", fgColor="FCE4D6"),
        "both_different": PatternFill("solid", fgColor="FFF2CC"),
    }

    def excel_value(value: Any):
        if isinstance(value, set):
            value = sorted(value)
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value.item() if hasattr(value, "item") else value

    def add_sheet(name: str, frame: pd.DataFrame) -> None:
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        columns = [str(column) for column in frame.columns]
        sheet.append(columns)
        for row in frame.itertuples(index=False, name=None):
            sheet.append([excel_value(value) for value in row])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
        sheet.freeze_panes = "A2"
        if columns:
            sheet.auto_filter.ref = sheet.dimensions
        sample_end = min(sheet.max_row, 501)
        for index, column in enumerate(columns, 1):
            values = [column]
            values.extend(
                "" if sheet.cell(row, index).value is None else str(sheet.cell(row, index).value)
                for row in range(2, sample_end + 1)
            )
            width = min(max(max(map(len, values)) + 2, 10), 42)
            sheet.column_dimensions[get_column_letter(index)].width = width
            if column.endswith("_rate") and sheet.max_row >= 2:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, index).number_format = "0.000%"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        if "v2_vs_improved_case" in columns and sheet.max_row >= 2:
            column_index = columns.index("v2_vs_improved_case") + 1
            column_letter = get_column_letter(column_index)
            target = f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            for case, fill in change_fills.items():
                sheet.conditional_formatting.add(
                    target,
                    FormulaRule(
                        formula=[f'${column_letter}2="{case}"'],
                        fill=fill,
                    ),
                )

    summary_frame = _summary_rows(summary)
    add_sheet("Summary", summary_frame)
    summary_sheet = workbook["Summary"]
    for row in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row, 1).value in {"scope", "note", "comparison"}:
            summary_sheet.cell(row, 1).fill = section_fill
    add_sheet("ByPlant", by_plant)
    add_sheet("ByDevice", by_device)
    add_sheet("ConfirmedEvents", confirmed_events)
    add_sheet("V2RawChanges", v2_raw_changes)
    add_sheet("V2Changes", v2_changes)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-root", required=True)
    parser.add_argument("--plant-id", action="append", default=[])
    parser.add_argument("--timezone", default="Asia/Shanghai")
    parser.add_argument("--interval-minutes", type=int, default=5)
    parser.add_argument("--output-directory", required=True)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    input_root = Path(args.input_root)
    if args.plant_id:
        plant_directories = [input_root / f"plant_id={plant}" for plant in args.plant_id]
    else:
        plant_directories = sorted(input_root.glob("plant_id=*"))
    if not plant_directories:
        raise FileNotFoundError(f"No plant_id=* replay directories under {input_root}")
    missing = [str(path) for path in plant_directories if not path.is_dir()]
    if missing:
        raise FileNotFoundError(f"Missing replay directories: {missing}")
    contracts = _load_contracts(plant_directories)
    device_coverage = _load_device_coverage(plant_directories)
    if contracts[0].get("timezone") != args.timezone:
        raise ValueError(
            "--timezone does not match the replay contract: "
            f"{contracts[0].get('timezone')}"
        )
    if int(contracts[0].get("interval_minutes")) != args.interval_minutes:
        raise ValueError(
            "--interval-minutes does not match the replay contract: "
            f"{contracts[0].get('interval_minutes')}"
        )

    output = Path(args.output_directory)
    _prepare_output(output, overwrite=args.overwrite)

    raw_points = _read_layer(plant_directories, signal="raw")
    final_points = _read_layer(plant_directories, signal="final")
    confirmed_points = _read_confirmed_candidate_points(plant_directories)
    confirmed_events = build_confirmed_events(
        confirmed_points,
        timezone=args.timezone,
        interval_minutes=args.interval_minutes,
    )
    v2_changes = confirmed_events[
        confirmed_events["v2_vs_improved_case"].isin(
            ["improved_only", "v2_only", "both_different"]
        )
    ].reset_index(drop=True)
    if not v2_changes.empty:
        v2_changes["row"] = range(1, len(v2_changes) + 1)
    by_plant = _scope_summary(
        raw_points,
        final_points,
        confirmed_events,
        keys=["plant_id"],
        universe=(
            device_coverage.groupby("plant_id", observed=True)
            .agg(
                devices=("device_no", "nunique"),
                device_time_rows=("device_time_rows", "sum"),
                first_time_local=("first_time_local", "min"),
                last_time_local=("last_time_local", "max"),
            )
            .reset_index()
        ),
    )
    by_device = _scope_summary(
        raw_points,
        final_points,
        confirmed_events,
        keys=["plant_id", "device_no"],
        universe=device_coverage,
    )

    for frame in (raw_points, final_points, confirmed_points):
        frame["event_time_local"] = (
            frame["event_time"]
            .dt.tz_convert(args.timezone)
            .dt.strftime("%Y-%m-%d %H:%M:%S")
        )
        frame.drop(columns="event_time", inplace=True)
        frame.insert(0, "row", range(1, len(frame) + 1))
    v2_raw_changes = raw_points[
        raw_points["v2_vs_improved_case"].isin(
            ["improved_only", "v2_only", "both_different"]
        )
    ].reset_index(drop=True)
    if not v2_raw_changes.empty:
        v2_raw_changes["row"] = range(1, len(v2_raw_changes) + 1)

    paths = {
        "raw_candidate_points": output / "pvlof_58d_raw_candidate_points.csv",
        "final_alert_points": output / "pvlof_58d_final_alert_points.csv",
        "confirmed_candidate_points": output
        / "pvlof_58d_confirmed_candidate_points.csv",
        "confirmed_events": output / "pvlof_58d_confirmed_events.csv",
        "v2_changes": output / "pvlof_58d_v2_changes.csv",
        "v2_raw_changes": output / "pvlof_58d_v2_raw_candidate_changes.csv",
        "by_plant": output / "pvlof_58d_by_plant.csv",
        "by_device": output / "pvlof_58d_by_device.csv",
        "workbook": output / "pvlof_58d_version_comparison.xlsx",
    }
    raw_points.to_csv(paths["raw_candidate_points"], index=False, encoding="utf-8-sig")
    final_points.to_csv(paths["final_alert_points"], index=False, encoding="utf-8-sig")
    confirmed_points.to_csv(
        paths["confirmed_candidate_points"], index=False, encoding="utf-8-sig"
    )
    confirmed_events.to_csv(paths["confirmed_events"], index=False, encoding="utf-8-sig")
    v2_changes.to_csv(paths["v2_changes"], index=False, encoding="utf-8-sig")
    v2_raw_changes.to_csv(
        paths["v2_raw_changes"], index=False, encoding="utf-8-sig"
    )
    by_plant.to_csv(paths["by_plant"], index=False, encoding="utf-8-sig")
    by_device.to_csv(paths["by_device"], index=False, encoding="utf-8-sig")

    version_summary = _global_version_summary(
        raw_points,
        final_points,
        confirmed_points,
        confirmed_events,
        evaluated_device_times=int(device_coverage["device_time_rows"].sum()),
    )
    comparison_summary = {
        "raw_candidate_rows": int(len(raw_points)),
        "final_alert_rows": int(len(final_points)),
        "confirmed_candidate_rows": int(len(confirmed_points)),
        "confirmed_events": int(len(confirmed_events)),
        "event_cases": confirmed_events["comparison_case"].value_counts().to_dict(),
        "v2_vs_improved_event_cases": confirmed_events[
            "v2_vs_improved_case"
        ].value_counts().to_dict(),
        "v2_changed_events": int(len(v2_changes)),
        "v2_changed_raw_candidate_rows": int(len(v2_raw_changes)),
    }
    summary = {
        "scope": {
            "plants": [str(contract.get("plant_id")) for contract in contracts],
            "devices": int(len(device_coverage)),
            "evaluated_device_times": int(device_coverage["device_time_rows"].sum()),
            "start_local_inclusive": contracts[0].get("start_local_inclusive"),
            "end_local_exclusive": contracts[0].get("end_local_exclusive"),
            "timezone": args.timezone,
            "interval_minutes": args.interval_minutes,
            "alarm_device_filter": False,
        },
        "versions": version_summary,
        "comparison": comparison_summary,
        "inputs": [str(path) for path in plant_directories],
        "outputs": {name: str(path) for name, path in paths.items()},
    }
    _write_workbook(
        paths["workbook"],
        summary=summary,
        by_plant=by_plant,
        by_device=by_device,
        confirmed_events=confirmed_events,
        v2_raw_changes=v2_raw_changes,
        v2_changes=v2_changes,
    )
    summary_path = output / "summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
