"""Build the manual review workbook for historical, replayed and improved PVLOF."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd


V16_HEADER = "PVLOF_V1_6_HYBRID_GATE"
V16_REPLAY_HEADER = "PVLOF_V1_6_V3_REPLAY"
V17_HEADER = "PVLOF_V1_7_PENALIZED_SEGMENTATION"
V17_IMPROVED_HEADER = "PVLOF_V1_7_IMPROVED"
V16_FINAL_HEADER = "PVLOF_V1_6_V3_FINAL_ALERT"
V17_FINAL_HEADER = "PVLOF_V1_7_FINAL_ALERT"
V17_IMPROVED_FINAL_HEADER = "PVLOF_V1_7_IMPROVED_FINAL_ALERT"
KEY_HEADERS = ("plant_id", "device_no", "event_time_local")
ALERT_PATTERN = re.compile(r"^(?:FALSE|\d{2}(?:,\d{2})*)$")


DEFINITIONS = {
    V16_REPLAY_HEADER: (
        ("pvlof_v16_raw_anomaly", "collective_raw_alert", "pvlof_v16_alert"),
        "pvlof_v16_alert",
    ),
    V17_HEADER: (
        ("pvlof_v17_raw_anomaly", "pvlof_v17_segmentation_raw_candidate", "pvlof_v17_alert"),
        "pvlof_v17_alert",
    ),
    V17_IMPROVED_HEADER: (
        (
            "pvlof_v17_improved_raw_anomaly",
            "pvlof_v17_improved_segmentation_raw_candidate",
            "pvlof_v17_improved_alert",
        ),
        "pvlof_v17_improved_alert",
    ),
}


def _normalise_plant(value: Any) -> int:
    return int(float(str(value).strip()))


def _normalise_device(value: Any) -> str:
    result = str(value).strip()
    if not result:
        raise ValueError("device_no is blank")
    return result


def _time_text(value: Any, timezone: str) -> str:
    timestamp = pd.Timestamp(value)
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert(timezone).tz_localize(None)
    return timestamp.strftime("%Y-%m-%d %H:%M:%S")


def _numbers(values: pd.Series) -> str:
    numbers = sorted({int(value) for value in values})
    return ",".join(f"{number:02d}" for number in numbers)


def _prediction_lookups(path, timezone, review_columns, final_column):
    source = pd.read_parquet(path)
    required = {"plant_id", "device_no", "event_time", "string_no", "string_current", final_column}
    missing = sorted(required - set(source.columns))
    if missing:
        raise ValueError(f"Prediction file {path} is missing columns: {missing}")
    available_review = [column for column in review_columns if column in source]
    if not available_review:
        raise ValueError(f"Prediction file {path} has no review evidence columns")

    source = source.copy()
    source["plant_id"] = pd.to_numeric(source["plant_id"], errors="raise").astype(int)
    source["device_no"] = source["device_no"].astype(str).str.strip()
    source["event_time_local"] = (
        pd.to_datetime(source["event_time"], errors="raise", utc=True)
        .dt.tz_convert(timezone)
        .dt.strftime("%Y-%m-%d %H:%M:%S")
    )
    source["string_no"] = pd.to_numeric(source["string_no"], errors="raise").astype(int)
    source["string_current"] = pd.to_numeric(source["string_current"], errors="coerce")
    source["_review"] = False
    for column in available_review:
        source["_review"] |= source[column].fillna(False).astype(bool)
    source["_final"] = source[final_column].fillna(False).astype(bool)

    key_columns = ["plant_id", "device_no", "event_time_local"]
    keys = {
        (int(row.plant_id), str(row.device_no), str(row.event_time_local))
        for row in source[key_columns].drop_duplicates().itertuples(index=False)
    }

    def lookup(flag):
        selected = source[source[flag] & source["string_current"].gt(0)]
        grouped = selected.groupby(key_columns, observed=True)["string_no"].agg(_numbers)
        return {tuple(index): value for index, value in grouped.items()}

    return lookup("_review"), lookup("_final"), keys, {
        "path": str(path),
        "rows": int(len(source)),
        "device_time_points": int(len(keys)),
        "review_columns": available_review,
        "final_column": final_column,
    }


def _headers(worksheet):
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _copy_style(worksheet, source_column, target_column):
    source_letter = worksheet.cell(1, source_column).column_letter
    target_letter = worksheet.cell(1, target_column).column_letter
    for row in range(1, worksheet.max_row + 1):
        source = worksheet.cell(row, source_column)
        target = worksheet.cell(row, target_column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    source_dimension = worksheet.column_dimensions[source_letter]
    target_dimension = worksheet.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = source_dimension.hidden


def _adjust_freeze_panes(worksheet, inserted_column):
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    frozen = worksheet.freeze_panes
    coordinate = frozen.coordinate if hasattr(frozen, "coordinate") else frozen
    if not coordinate:
        return
    column_text, row = coordinate_from_string(str(coordinate))
    column = column_index_from_string(column_text)
    if column >= inserted_column:
        column += 1
    worksheet.freeze_panes = f"{get_column_letter(column)}{row}"


def _ensure_after(worksheet, after_header, header):
    headers = _headers(worksheet)
    if header in headers:
        return headers[header]
    if after_header not in headers:
        raise ValueError(f"Required preceding column is missing: {after_header}")
    target = headers[after_header] + 1
    worksheet.insert_cols(target, 1)
    _copy_style(worksheet, target - 1, target)
    _adjust_freeze_panes(worksheet, target)
    worksheet.cell(1, target).value = header
    return target


def build_review_workbook(
    input_workbook,
    v16_predictions,
    v17_predictions,
    improved_predictions,
    output_workbook,
    report_path,
    *,
    timezone="Asia/Shanghai",
    sheet_name=None,
):
    from openpyxl import load_workbook

    input_path = Path(input_workbook).resolve()
    output_path = Path(output_workbook).resolve()
    if input_path == output_path:
        raise ValueError("Output workbook must differ from input workbook")

    sources = {
        V16_REPLAY_HEADER: (v16_predictions, V16_FINAL_HEADER),
        V17_HEADER: (v17_predictions, V17_FINAL_HEADER),
        V17_IMPROVED_HEADER: (improved_predictions, V17_IMPROVED_FINAL_HEADER),
    }
    lookups, prediction_reports = {}, {}
    key_sets = {}
    for header, (path, final_header) in sources.items():
        review_columns, final_column = DEFINITIONS[header]
        review, final, keys, report = _prediction_lookups(
            path, timezone, review_columns, final_column
        )
        lookups[header] = review
        lookups[final_header] = final
        key_sets[header] = keys
        prediction_reports[header] = report

    workbook = load_workbook(input_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = _headers(worksheet)
    missing = sorted({"Baseline", V16_HEADER, *KEY_HEADERS} - set(headers))
    if missing:
        raise ValueError(f"Review workbook is missing headers: {missing}")

    manual_values = [worksheet.cell(row, 1).value for row in range(1, worksheet.max_row + 1)]
    baseline_values = [
        worksheet.cell(row, headers["Baseline"]).value
        for row in range(1, worksheet.max_row + 1)
    ]
    historical_v16_values = [
        worksheet.cell(row, headers[V16_HEADER]).value
        for row in range(1, worksheet.max_row + 1)
    ]

    _ensure_after(worksheet, V16_HEADER, V16_REPLAY_HEADER)
    _ensure_after(worksheet, V16_REPLAY_HEADER, V17_HEADER)
    _ensure_after(worksheet, V17_HEADER, V17_IMPROVED_HEADER)
    _ensure_after(worksheet, V17_IMPROVED_HEADER, V16_FINAL_HEADER)
    _ensure_after(worksheet, V16_FINAL_HEADER, V17_FINAL_HEADER)
    _ensure_after(worksheet, V17_FINAL_HEADER, V17_IMPROVED_FINAL_HEADER)
    headers = _headers(worksheet)

    workbook_keys = []
    for row in range(2, worksheet.max_row + 1):
        workbook_keys.append((
            _normalise_plant(worksheet.cell(row, headers["plant_id"]).value),
            _normalise_device(worksheet.cell(row, headers["device_no"]).value),
            _time_text(worksheet.cell(row, headers["event_time_local"]).value, timezone),
        ))
    workbook_key_set = set(workbook_keys)
    coverage = {}
    for header, keys in key_sets.items():
        missing_keys = workbook_key_set - keys
        coverage[header] = {"missing": len(missing_keys)}
        if missing_keys:
            raise ValueError(
                f"{header} does not cover every workbook point: missing={len(missing_keys)}, "
                f"examples={list(sorted(missing_keys))[:5]}"
            )

    result_headers = [
        V16_REPLAY_HEADER, V17_HEADER, V17_IMPROVED_HEADER,
        V16_FINAL_HEADER, V17_FINAL_HEADER, V17_IMPROVED_FINAL_HEADER,
    ]
    counts = {header: 0 for header in result_headers}
    for row, key in enumerate(workbook_keys, start=2):
        for header in result_headers:
            value = lookups[header].get(key, "FALSE")
            if not ALERT_PATTERN.fullmatch(value):
                raise ValueError(f"Invalid review value for {header}, {key}: {value}")
            worksheet.cell(row, headers[header]).value = value
            counts[header] += int(value != "FALSE")

    if manual_values != [worksheet.cell(row, 1).value for row in range(1, worksheet.max_row + 1)]:
        raise AssertionError("Manual label column changed during export")
    if baseline_values != [worksheet.cell(row, headers["Baseline"]).value for row in range(1, worksheet.max_row + 1)]:
        raise AssertionError("Baseline column changed during export")
    if historical_v16_values != [worksheet.cell(row, headers[V16_HEADER]).value for row in range(1, worksheet.max_row + 1)]:
        raise AssertionError("Historical v1.6 column changed during export")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    report = {
        "input_workbook": str(input_path),
        "output_workbook": str(output_path),
        "sheet": worksheet.title,
        "rows": int(worksheet.max_row - 1),
        "manual_label_preserved": True,
        "baseline_preserved": True,
        "historical_v16_preserved": True,
        "coverage": coverage,
        "non_false_rows": counts,
        "definitions": {
            "review_columns": "raw candidate OR final alert; positive-current strings",
            "final_columns": "final alert only; positive-current strings",
        },
        "predictions": prediction_reports,
    }
    destination = Path(report_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-workbook", required=True)
    parser.add_argument("--v16-predictions", required=True)
    parser.add_argument("--v17-predictions", required=True)
    parser.add_argument("--improved-predictions", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--timezone", default="Asia/Shanghai")
    parser.add_argument("--sheet")
    args = parser.parse_args()
    build_review_workbook(
        args.input_workbook,
        args.v16_predictions,
        args.v17_predictions,
        args.improved_predictions,
        args.output,
        args.report,
        timezone=args.timezone,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
