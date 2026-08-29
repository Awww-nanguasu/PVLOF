"""Add PVLOF v1.7 improved V2 results without replacing prior review columns."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


from scripts.build_pvlof_v17_improved_review_workbook import (  # noqa: E402
    ALERT_PATTERN,
    KEY_HEADERS,
    V17_IMPROVED_FINAL_HEADER,
    V17_IMPROVED_HEADER,
    _ensure_after,
    _headers,
    _normalise_device,
    _normalise_plant,
    _prediction_lookups,
    _time_text,
)


V17_IMPROVED_V2_HEADER = "PVLOF_V1_7_IMPROVED_V2"
V17_IMPROVED_V2_FINAL_HEADER = "PVLOF_V1_7_IMPROVED_V2_FINAL_ALERT"
V2_REVIEW_COLUMNS = (
    "pvlof_v17_improved_raw_anomaly",
    "pvlof_v17_improved_segmentation_raw_candidate",
    "pvlof_v17_improved_alert",
)
V2_FINAL_COLUMN = "pvlof_v17_improved_alert"


def _column_values(worksheet, column):
    return [
        worksheet.cell(row, column).value
        for row in range(1, worksheet.max_row + 1)
    ]


def add_v2_results(
    input_workbook,
    improved_v2_predictions,
    output_workbook,
    report_path,
    *,
    timezone="Asia/Shanghai",
    sheet_name=None,
):
    from openpyxl import load_workbook

    input_path = Path(input_workbook).resolve()
    output_path = Path(output_workbook).resolve()
    if input_path == output_path:
        raise ValueError("Output workbook must differ from input workbook")

    review_lookup, final_lookup, prediction_keys, prediction_report = (
        _prediction_lookups(
            improved_v2_predictions,
            timezone,
            V2_REVIEW_COLUMNS,
            V2_FINAL_COLUMN,
        )
    )

    workbook = load_workbook(input_path)
    original_sheet_names = list(workbook.sheetnames)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    headers_before = _headers(worksheet)
    required = {
        "Baseline",
        V17_IMPROVED_HEADER,
        V17_IMPROVED_FINAL_HEADER,
        *KEY_HEADERS,
    }
    missing = sorted(required - set(headers_before))
    if missing:
        raise ValueError(f"Review workbook is missing headers: {missing}")

    manual_values = _column_values(worksheet, 1)
    preserved_headers = [
        header
        for header in headers_before
        if header not in {V17_IMPROVED_V2_HEADER, V17_IMPROVED_V2_FINAL_HEADER}
    ]
    preserved_values = {
        header: _column_values(worksheet, headers_before[header])
        for header in preserved_headers
    }

    _ensure_after(
        worksheet,
        V17_IMPROVED_HEADER,
        V17_IMPROVED_V2_HEADER,
    )
    _ensure_after(
        worksheet,
        V17_IMPROVED_FINAL_HEADER,
        V17_IMPROVED_V2_FINAL_HEADER,
    )
    headers = _headers(worksheet)

    workbook_keys = []
    for row in range(2, worksheet.max_row + 1):
        workbook_keys.append((
            _normalise_plant(worksheet.cell(row, headers["plant_id"]).value),
            _normalise_device(worksheet.cell(row, headers["device_no"]).value),
            _time_text(
                worksheet.cell(row, headers["event_time_local"]).value,
                timezone,
            ),
        ))
    workbook_key_set = set(workbook_keys)
    missing_keys = workbook_key_set - prediction_keys
    if missing_keys:
        raise ValueError(
            f"{V17_IMPROVED_V2_HEADER} does not cover every workbook point: "
            f"missing={len(missing_keys)}, examples={list(sorted(missing_keys))[:5]}"
        )

    raw_count = 0
    final_count = 0
    for row, key in enumerate(workbook_keys, start=2):
        raw_value = review_lookup.get(key, "FALSE")
        final_value = final_lookup.get(key, "FALSE")
        if not ALERT_PATTERN.fullmatch(raw_value):
            raise ValueError(f"Invalid V2 review value for {key}: {raw_value}")
        if not ALERT_PATTERN.fullmatch(final_value):
            raise ValueError(f"Invalid V2 final value for {key}: {final_value}")
        worksheet.cell(row, headers[V17_IMPROVED_V2_HEADER]).value = raw_value
        worksheet.cell(row, headers[V17_IMPROVED_V2_FINAL_HEADER]).value = final_value
        raw_count += int(raw_value != "FALSE")
        final_count += int(final_value != "FALSE")

    if manual_values != _column_values(worksheet, 1):
        raise AssertionError("Manual label column changed during V2 export")
    headers_after = _headers(worksheet)
    for header, expected in preserved_values.items():
        actual = _column_values(worksheet, headers_after[header])
        if expected != actual:
            raise AssertionError(f"Existing column changed during V2 export: {header}")
    if original_sheet_names != list(workbook.sheetnames):
        raise AssertionError("Workbook sheet names changed during V2 export")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    report = {
        "input_workbook": str(input_path),
        "output_workbook": str(output_path),
        "sheet": worksheet.title,
        "rows": int(worksheet.max_row - 1),
        "manual_label_preserved": True,
        "existing_columns_preserved": True,
        "sheet_names_preserved": True,
        "coverage": {
            "workbook_points": len(workbook_key_set),
            "missing": 0,
        },
        "non_false_rows": {
            V17_IMPROVED_V2_HEADER: raw_count,
            V17_IMPROVED_V2_FINAL_HEADER: final_count,
        },
        "definitions": {
            V17_IMPROVED_V2_HEADER: "raw candidate OR final alert; positive-current strings",
            V17_IMPROVED_V2_FINAL_HEADER: "final alert only; positive-current strings",
        },
        "predictions": prediction_report,
    }
    destination = Path(report_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-workbook", required=True)
    parser.add_argument("--improved-v2-predictions", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--timezone", default="Asia/Shanghai")
    parser.add_argument("--sheet")
    args = parser.parse_args()
    add_v2_results(
        args.input_workbook,
        args.improved_v2_predictions,
        args.output,
        args.report,
        timezone=args.timezone,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()


__all__ = [
    "V17_IMPROVED_V2_FINAL_HEADER",
    "V17_IMPROVED_V2_HEADER",
    "add_v2_results",
]
