"""Add PVLOF v1.7 string results to the existing v1.6 review workbook."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd


V16_HEADER = "PVLOF_V1_6_HYBRID_GATE"
V17_HEADER = "PVLOF_V1_7_PENALIZED_SEGMENTATION"
WORKBOOK_KEY_HEADERS = ("plant_id", "device_no", "event_time_local")
PREDICTION_KEYS = ("plant_id", "device_no", "event_time")
V17_EVIDENCE_COLUMNS = (
    "pvlof_v17_raw_anomaly",
    "pvlof_v17_segmentation_raw_candidate",
    "pvlof_v17_alert",
)
ALERT_TEXT_PATTERN = re.compile(r"^(?:FALSE|\d{2}(?:,\d{2})*)$")


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required. Run this script with the local Anaconda python "
            "environment that provides openpyxl."
        ) from error
    return (
        load_workbook,
        get_column_letter,
        coordinate_from_string,
        column_index_from_string,
    )


def _normalize_plant(value: Any) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise ValueError("plant_id is blank")
    return int(float(str(value).strip()))


def _normalize_device(value: Any) -> str:
    result = str(value).strip()
    if not result:
        raise ValueError("device_no is blank")
    return result


def _workbook_time_text(value: Any, timezone: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise ValueError("event_time_local is blank")
    timestamp = pd.Timestamp(value)
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert(timezone).tz_localize(None)
    return timestamp.strftime("%Y-%m-%d %H:%M:%S")


def _prediction_lookup(
    path: str | Path,
    timezone: str,
) -> tuple[dict[tuple[int, str, str], str], set[tuple[int, str, str]], dict[str, Any]]:
    source = pd.read_parquet(path)
    required = {*PREDICTION_KEYS, "string_no", "string_current"}
    missing = sorted(required - set(source.columns))
    if missing:
        raise ValueError(f"V1.7 predictions are missing columns: {missing}")
    evidence_columns = [
        column for column in V17_EVIDENCE_COLUMNS if column in source.columns
    ]
    if not evidence_columns:
        raise ValueError(
            "V1.7 predictions contain none of the supported evidence columns: "
            f"{list(V17_EVIDENCE_COLUMNS)}"
        )

    source = source.copy()
    source["plant_id"] = pd.to_numeric(
        source["plant_id"], errors="raise"
    ).astype(int)
    source["device_no"] = source["device_no"].astype(str).str.strip()
    source["event_time_local"] = (
        pd.to_datetime(source["event_time"], errors="raise", utc=True)
        .dt.tz_convert(timezone)
        .dt.strftime("%Y-%m-%d %H:%M:%S")
    )
    source["string_no"] = pd.to_numeric(
        source["string_no"], errors="raise"
    ).astype(int)
    source["string_current"] = pd.to_numeric(
        source["string_current"], errors="coerce"
    )
    source["_v17_review_evidence"] = False
    for column in evidence_columns:
        source["_v17_review_evidence"] |= source[column].fillna(False).astype(bool)

    local_keys = ["plant_id", "device_no", "event_time_local"]
    prediction_keys = {
        (int(row.plant_id), str(row.device_no), str(row.event_time_local))
        for row in source[local_keys].drop_duplicates().itertuples(index=False)
    }
    alerted = source[
        source["_v17_review_evidence"] & source["string_current"].gt(0)
    ]
    grouped = (
        alerted.groupby(local_keys, observed=True)["string_no"]
        .agg(lambda values: ",".join(f"{value:02d}" for value in sorted(set(values))))
        .reset_index()
    )
    lookup = {
        (int(row.plant_id), str(row.device_no), str(row.event_time_local)): str(
            row.string_no
        )
        for row in grouped.itertuples(index=False)
    }
    report = {
        "prediction_file": str(path),
        "prediction_string_rows": int(len(source)),
        "prediction_device_time_points": int(len(prediction_keys)),
        "prediction_alert_device_time_points": int(len(lookup)),
        "evidence_columns": evidence_columns,
        "definition": (
            "positive-current strings where any available V1.7 raw candidate or "
            "final alert field is true"
        ),
    }
    return lookup, prediction_keys, report


def _copy_column_style(worksheet, source_column: int, target_column: int) -> None:
    for row in range(1, worksheet.max_row + 1):
        source = worksheet.cell(row=row, column=source_column)
        target = worksheet.cell(row=row, column=target_column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def add_v17_column(
    input_workbook: str | Path,
    v17_predictions: str | Path,
    output_workbook: str | Path,
    report_path: str | Path,
    *,
    timezone: str = "Asia/Shanghai",
    sheet_name: str | None = None,
    require_full_coverage: bool = True,
) -> dict[str, Any]:
    input_path = Path(input_workbook).resolve()
    output_path = Path(output_workbook).resolve()
    if input_path == output_path:
        raise ValueError("Output workbook must differ from the input workbook")

    (
        load_workbook,
        get_column_letter,
        coordinate_from_string,
        column_index_from_string,
    ) = _load_openpyxl()
    lookup, prediction_keys, prediction_report = _prediction_lookup(
        v17_predictions, timezone
    )

    workbook = load_workbook(input_path)
    if sheet_name is None:
        worksheet = workbook.active
    else:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]

    headers = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    missing_headers = sorted(
        {V16_HEADER, *WORKBOOK_KEY_HEADERS} - set(headers)
    )
    if missing_headers:
        raise ValueError(f"Review workbook is missing headers: {missing_headers}")

    if V17_HEADER in headers:
        v17_column = headers[V17_HEADER]
        inserted = False
    else:
        v16_column = headers[V16_HEADER]
        v17_column = v16_column + 1
        frozen = worksheet.freeze_panes
        frozen_coordinate = frozen.coordinate if hasattr(frozen, "coordinate") else frozen
        worksheet.insert_cols(v17_column, 1)
        _copy_column_style(worksheet, v16_column, v17_column)

        source_letter = get_column_letter(v16_column)
        target_letter = get_column_letter(v17_column)
        source_dimension = worksheet.column_dimensions[source_letter]
        target_dimension = worksheet.column_dimensions[target_letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel

        if frozen_coordinate:
            frozen_column_text, frozen_row = coordinate_from_string(
                str(frozen_coordinate)
            )
            frozen_column = column_index_from_string(frozen_column_text)
            if frozen_column >= v17_column:
                frozen_column += 1
            worksheet.freeze_panes = (
                f"{get_column_letter(frozen_column)}{frozen_row}"
            )
        inserted = True

    worksheet.cell(row=1, column=v17_column).value = V17_HEADER
    headers = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    workbook_keys: set[tuple[int, str, str]] = set()
    row_keys: list[tuple[int, str, str]] = []
    for row in range(2, worksheet.max_row + 1):
        key = (
            _normalize_plant(worksheet.cell(row, headers["plant_id"]).value),
            _normalize_device(worksheet.cell(row, headers["device_no"]).value),
            _workbook_time_text(
                worksheet.cell(row, headers["event_time_local"]).value,
                timezone,
            ),
        )
        workbook_keys.add(key)
        row_keys.append(key)

    missing_prediction_keys = sorted(workbook_keys - prediction_keys)
    if require_full_coverage and missing_prediction_keys:
        examples = [
            {
                "plant_id": key[0],
                "device_no": key[1],
                "event_time_local": key[2],
            }
            for key in missing_prediction_keys[:10]
        ]
        raise ValueError(
            "V1.7 predictions do not cover every workbook device/time point: "
            f"missing={len(missing_prediction_keys)}, examples={examples}. "
            "Regenerate V1.7 with the same floor-aligned alarm points before export."
        )

    output_alert_rows = 0
    for row, key in enumerate(row_keys, start=2):
        value = lookup.get(key, "FALSE")
        if not ALERT_TEXT_PATTERN.fullmatch(value):
            raise ValueError(f"Invalid V1.7 review value for {key}: {value}")
        worksheet.cell(row=row, column=v17_column).value = value
        output_alert_rows += int(value != "FALSE")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    report = {
        "input_workbook": str(input_path),
        "output_workbook": str(output_path),
        "sheet": worksheet.title,
        "v17_column": V17_HEADER,
        "column_inserted": inserted,
        "timezone": timezone,
        "workbook_rows": int(worksheet.max_row - 1),
        "workbook_unique_device_time_points": int(len(workbook_keys)),
        "missing_prediction_device_time_points": int(len(missing_prediction_keys)),
        "v17_alert_rows": int(output_alert_rows),
        "prediction": prediction_report,
    }
    report_destination = Path(report_path)
    report_destination.parent.mkdir(parents=True, exist_ok=True)
    report_destination.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-workbook",
        default="pvlof_v16_vs_baseline_floor_raise.xlsx",
    )
    parser.add_argument(
        "--v17-predictions",
        default=(
            "artifacts/models/pvlof_v17/alarm_windows_v2_floor_raise/"
            "pvlof_v17_alarm_points.parquet"
        ),
    )
    parser.add_argument(
        "--output",
        default="pvlof_v17_vs_baseline_floor_raise.xlsx",
    )
    parser.add_argument(
        "--report",
        default=(
            "artifacts/reports/pvlof_v17_vs_baseline_floor_raise/summary.json"
        ),
    )
    parser.add_argument("--timezone", default="Asia/Shanghai")
    parser.add_argument("--sheet")
    parser.add_argument(
        "--allow-missing",
        action="store_true",
        help="Fill uncovered workbook points with FALSE instead of failing",
    )
    args = parser.parse_args()
    add_v17_column(
        args.input_workbook,
        args.v17_predictions,
        args.output,
        args.report,
        timezone=args.timezone,
        sheet_name=args.sheet,
        require_full_coverage=not args.allow_missing,
    )


if __name__ == "__main__":
    main()
